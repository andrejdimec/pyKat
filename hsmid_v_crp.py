import arcpy
from arcpy import env
from openpyxl import load_workbook

import vars
import time
from logger import Logger
from comm import Comm

black = vars.black
red = vars.darkRed
green = vars.darkGreen
blue = vars.darkBlue
err = vars.red

hs_fc = r"RPE\hisne_stevilke"
na_fc = r"RPE\NA_2024"
ul_fc = r"RPE\UL_2024"


def stop(st):
    end_time = time.time()
    elapsed_time = end_time - st
    m = int(elapsed_time / 60)
    s = int(elapsed_time % 60)
    return m, s


# najdi eid ulice
def eid_ulica(naziv_ulice, eid_obcine):
    vrni = None
    query = f"NAZIV = '{naziv_ulice}' AND EID_OBCINA='{eid_obcine}'"
    with arcpy.da.SearchCursor(ul_fc, ["EID_ULICA"], query) as cursor:
        for row in cursor:
            vrni = row[0]
    return vrni


# Najdi eid naselja
def eid_naselje(naziv_naselja, eid_obcine):
    vrni = None
    query = f"NAZIV = '{naziv_naselja}' AND EID_OBCINA='{eid_obcine}'"
    with arcpy.da.SearchCursor(na_fc, ["EID_NASELJ"], query) as cursor:
        for row in cursor:
            vrni = row[0]
    return vrni


# najdi eid hišne številke
def eid_hs(eid_naselja, eid_ulice, hs, hsd):
    eidhs = None
    query = f"EID_NASELJE = '{eid_naselja}' AND EID_ULICA='{eid_ulice}' AND HS_STEVILKA={hs} AND HS_DODATEK='{hsd}'"

    if eid_ulice is None and hsd is None:
        query = f"EID_NASELJE = '{eid_naselja}' AND HS_STEVILKA={hs}"

    elif eid_ulice is None and hsd is not None:
        query = (
            f"EID_NASELJE = '{eid_naselja}' AND HS_STEVILKA={hs} AND HS_DODATEK='{hsd}'"
        )
    elif eid_ulice is not None and hsd is None:
        query = f"EID_NASELJE = '{eid_naselja}' AND EID_ULICA='{eid_ulice}' AND HS_STEVILKA={hs}"

    with arcpy.da.SearchCursor(hs_fc, ["EID_HISNA_STEVILKA"], query) as cursor:
        for row in cursor:
            eidhs = row[0]
        if eidhs:
            hsmid = eidhs[9 : len(eidhs) - 1]

    return eidhs, hsmid


class HsmidCrp(Comm):
    def __init__(self, comm=Comm()):
        super(HsmidCrp, self).__init__()
        self.comm = comm
        self.logger = Logger(comm=self.comm)

    def hsmid_crp(self, filename):
        env.workspace = vars.wkspace
        # self.log("Workspace: " + str(env.workspace))
        # self.log(f"Arcgis datoteka: {hs_fc}")

        start_time = time.time()
        stalnih = 0
        zacasnih = 0
        stalnih_zacasnih = 0
        nikjer = 0

        self.logc("V CRP dodaj EID in HSMID za stalne in začasne prebivalce.", green)
        # filename = "d:/podatki/crp-01-07-2024.xlsx"
        self.presledek()
        self.log(filename)
        workbook = load_workbook(filename)
        sheet = workbook.active
        rows = sheet.max_row
        self.log("Število prebivalcev v CRP:" + str(rows))
        self.presledek()
        self.log("Prenašam...")
        #
        # preveri če kolone za hsmid in hseid že obstajajo
        headers_to_delete = ["hsmid_st", "hseid_st", "hsmid_zc", "hseid_zc"]

        # Find the column indices based on the header names
        columns_to_delete = []
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[0].value in headers_to_delete:
                columns_to_delete.append(col[0].column)

        # Delete the columns if found
        for col_index in sorted(columns_to_delete, reverse=True):
            sheet.delete_cols(col_index)

        # Add the headers as the last columns
        last_col_index = sheet.max_column + 1
        for header in headers_to_delete:
            sheet.cell(row=1, column=last_col_index, value=header)
            last_col_index += 1

        workbook.save(filename)

        # Poišči indekse kolon, ki jih rabimo

        header = {cell.value: cell.column for cell in sheet[1]}

        st_obcina_col = header["Občina STPR"] - 1
        st_naselje_col = header["Naziv naselja STPR"] - 1
        st_ulica_col = header["Naziv ulice STPR"] - 1
        st_hs_col = header["Hišna št. STPR"] - 1
        st_hsd_col = header["Hišna št. STPR dodatek"] - 1
        zc_obcina_col = header["Občina ZCPR"] - 1
        zc_naselje_col = header["Naziv naselja ZCPR"] - 1
        zc_ulica_col = header["Naziv ulice ZCPR"] - 1
        zc_hs_col = header["Hišna št. ZCPR"] - 1
        zc_hsd_col = header["Hišna št. ZCPR dodatek"] - 1
        st_hsmid_col = header["hsmid_st"]
        zc_hsmid_col = header["hsmid_zc"]
        st_hseid_col = header["hseid_st"]
        zc_hseid_col = header["hseid_zc"]

        # Obdelaj celo Excel tabelo in vpiši EID za HS
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            stalni = False
            zacasni = False
            # Stalno prebivališče
            st_obcina_eid = "-1"
            st_obcina_id = row[st_obcina_col].value
            if st_obcina_id == vars.obcina_apace_id:
                st_obcina_eid = vars.obcina_apace_eid
            if st_obcina_id == vars.obcina_radgona_id:
                st_obcina_eid = vars.obcina_radgona_eid
            if st_obcina_id == vars.obcina_radenci_id:
                st_obcina_eid = vars.obcina_radenci_eid
            if st_obcina_id == vars.obcina_jurij_id:
                st_obcina_eid = vars.obcina_jurij_eid

            if st_obcina_eid in vars.obcine_eid:
                st_naselje = row[st_naselje_col].value
                st_ulica = row[st_ulica_col].value
                st_naselje_eid = eid_naselje(st_naselje, st_obcina_eid)
                st_hs = row[st_hs_col].value
                st_hsd = row[st_hsd_col].value
                if st_hsd.replace(" ", "") == "":
                    st_hsd = None
                st_ulica_eid = eid_ulica(st_ulica, st_obcina_eid)

                eidhs, hsmid = eid_hs(st_naselje_eid, st_ulica_eid, st_hs, st_hsd)
                if eidhs:
                    stalni = True
                    stalnih += 1

                sheet.cell(row=row[0].row, column=st_hseid_col, value=eidhs)
                sheet.cell(row=row[0].row, column=st_hsmid_col, value=hsmid)

            else:
                sheet.cell(row=row[0].row, column=st_hseid_col, value="")
                sheet.cell(row=row[0].row, column=st_hsmid_col, value="")

            # Začasno prebivališče
            zc_obcina_eid = "-1"
            zc_obcina_id = row[zc_obcina_col].value
            if zc_obcina_id == vars.obcina_apace_id:
                zc_obcina_eid = vars.obcina_apace_eid
            if zc_obcina_id == vars.obcina_radgona_id:
                zc_obcina_eid = vars.obcina_radgona_eid
            if zc_obcina_id == vars.obcina_radenci_id:
                zc_obcina_eid = vars.obcina_radenci_eid
            if zc_obcina_id == vars.obcina_jurij_id:
                zc_obcina_eid = vars.obcina_jurij_eid

            if zc_obcina_eid in vars.obcine_eid:
                zc_naselje = row[zc_naselje_col].value
                zc_ulica = row[zc_ulica_col].value
                zc_naselje_eid = eid_naselje(zc_naselje, zc_obcina_eid)
                zc_hs = row[zc_hs_col].value
                zc_hsd = row[zc_hsd_col].value
                if zc_hsd.replace(" ", "") == "":
                    zc_hsd = None
                zc_ulica_eid = eid_ulica(zc_ulica, zc_obcina_eid)

                eidhs, hsmid = eid_hs(zc_naselje_eid, zc_ulica_eid, zc_hs, zc_hsd)

                sheet.cell(row=row[0].row, column=zc_hseid_col, value=eidhs)
                sheet.cell(row=row[0].row, column=zc_hsmid_col, value=hsmid)

            else:
                sheet.cell(row=row[0].row, column=zc_hseid_col, value="")
                sheet.cell(row=row[0].row, column=zc_hsmid_col, value="")

            workbook.save(filename)

        self.presledek()
        minutes, seconds = stop(start_time)
        self.logc(f"Končano v {minutes} min {seconds} sek.", blue)

    def log(self, in_str):
        self.logger.izpisi(in_str, black)

    def logc(self, in_str, barva):
        self.logger.izpisi(in_str, barva)

    def presledek(self):
        self.log("\n")
