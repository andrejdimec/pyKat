import arcpy
from arcpy import env
import openpyxl
from openpyxl import load_workbook

import vars
import time
from logger import Logger
from comm import Comm

black = vars.black
red = vars.darkRed
green = vars.darkGreen
blue = vars.darkBlue
err = vars.red


def stop(st):
    end_time = time.time()
    elapsed_time = end_time - st
    m = int(elapsed_time / 60)
    s = int(elapsed_time % 60)
    return m, s


class PrebivalciHs(Comm):
    def __init__(self, comm=Comm()):
        super(PrebivalciHs, self).__init__()
        self.comm = comm
        self.logger = Logger(comm=self.comm)

    def prebivalci_hs(self):
        env.workspace = vars.wkspace
        self.log("Workspace: " + str(env.workspace))
        hs_fc = r"RPE\hisne_stevilke"
        self.log(f"Arcgis datoteka: {hs_fc}")
        start_time = time.time()

        self.logc("Vpiši število prebivalcev v datoteko s hišnimi številkami.", green)
        filename = "d:/podatki/crp-01-07-2024.xlsx"

        self.log("Excel datoteka: " + filename)
        workbook = load_workbook(filename)
        sheet = workbook.active
        crp_list = []
        hs_list = []

        header = {cell.value: cell.column for cell in sheet[1]}
        eid_st_col = header["hseid_st"]
        eid_zc_col = header["hseid_zc"]

        # Naloži vse potrebne kolone iz excela v list
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            eid_st = row[eid_st_col - 1].value
            eid_zc = row[eid_zc_col - 1].value
            crp_list.append([eid_st, eid_zc])

        self.log("Število oseb v CRP " + str(len(crp_list)))

        self.log(f"Zapisujem podatke o prebivalcih v ArcGis tabelo {hs_fc}")
        if arcpy.TestSchemaLock(hs_fc):
            fields = ["EID_HISNA_STEVILKA", "prebivalcev", "stalno", "zacasno"]
            with arcpy.da.UpdateCursor(hs_fc, fields) as cursor:
                for row_hs in cursor:
                    eid_hs = row_hs[0]
                    prebivalcev = 0
                    stalno = 0
                    zacasno = 0
                    for oseba in crp_list:
                        st_preb = oseba[0]
                        zc_preb = oseba[1]
                        if st_preb == eid_hs:
                            stalno += 1
                            prebivalcev += 1
                        if zc_preb == eid_hs:
                            zacasno += 1
                            prebivalcev += 1
                    row_hs[1] = prebivalcev
                    row_hs[2] = stalno
                    row_hs[3] = zacasno
                    cursor.updateRow(row_hs)
        else:
            self.logc(f"Error: datoteka {hs_list} je zaklenjena", err)
        self.presledek()
        minutes, seconds = stop(start_time)
        self.logc(f"Končano v {minutes} min {seconds} sek.", blue)

    def log(self, in_str):
        self.logger.izpisi(in_str, black)

    def logc(self, in_str, barva):
        self.logger.izpisi(in_str, barva)

    def presledek(self):
        self.log("\n")
