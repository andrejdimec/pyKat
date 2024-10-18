import arcpy
from arcpy import env

from PySide6.QtCore import Signal, QObject
from PySide6.QtWidgets import QMessageBox

from openpyxl import load_workbook

import vars
import time
from logger import Logger
from comm import Comm


black = vars.black
red = vars.darkRed
green = vars.darkGreen
blue = vars.darkBlue
err = vars.red

hs_fc = r"RPE\hisne_stevilke"
na_fc = r"RPE\NA_2024"
ul_fc = r"RPE\UL_2024"


class MsgBox(QMessageBox):
    def __init__(self, tekst, parent=None):
        super().__init__(parent)
        self.tekst = tekst

        self.setWindowTitle("Sporočilo.")
        self.setIcon(QMessageBox.Icon.Information)
        self.setText(tekst)
        self.setStandardButtons(QMessageBox.StandardButton.Yes)
        self.buttonY = self.button(QMessageBox.StandardButton.Yes)
        self.buttonY.setText("V redu")
        MsgBox.exec(self)


def stop(st):
    end_time = time.time()
    elapsed_time = end_time - st
    m = int(elapsed_time / 60)
    s = int(elapsed_time % 60)
    return m, s


class UvoziInfotim(Comm):
    progress = Signal(int)
    status = Signal(str)

    def __init__(self, comm=Comm()):
        super(UvoziInfotim, self).__init__()
        self.comm = comm
        self.logger = Logger(comm=self.comm)

    def uvoz(self, filename):
        print("Uvoz števcev Infotim")
        env.workspace = vars.wkspace
        # self.log("Workspace: " + str(env.workspace))
        # vodomer_fc = r"Vodovod\vodomeri_infotim"
        vodomer_fc = r"Odjemna_mesta\infotim_1"

        self.log(f"Arcgis datoteka: {hs_fc}")

        start_time = time.time()
        stalnih = 0
        zacasnih = 0
        stalnih_zacasnih = 0
        nikjer = 0

        self.log(filename)
        workbook = load_workbook(filename)
        sheet = workbook.active
        rows = sheet.max_row
        self.logc(f"Uvoz {rows} števcev Infotim", green)
        # filename = "d:/podatki/crp-01-07-2024.xlsx"
        self.presledek()
        seznam_info = []
        # Get the column indices for the specified headings
        header = {cell.value: cell.column for cell in sheet[1]}

        # Kolone v listi
        x_col = header["GPS Lat"]
        y_col = header["GPS Lon"]
        # x_str=str(header["GPS Lat"])
        # y_str=str(header["GPS Lon"])
        # x_str=x_str.replace(",", ".")
        # y_str=y_str.replace(",", ".")

        koda_om_col = header["Koda OM"]
        lokacija_col = header["Lokacija"]
        tip_col = header["Tip oddajnika"]
        dimenzija_col = header["Dimenzija"]
        naslov_col = header["Naslov"]
        lastnik_col = header["Lastnik"]
        seriska_col = header["Serijska številka"]
        rf_col = header["RF Naslov"]

        # Naloži vse potrebne kolone iz excela v list
        for row in sheet.iter_rows(min_row=2, max_row=rows):
            # x_str = str(row[x_col - 1].value).replace(".", ",")
            # y_str = str(row[y_col - 1].value).replace(".", ",")
            x = row[x_col - 1].value
            y = row[y_col - 1].value
            print("x: " + str(x) + " y: " + str(y))
            print("x-10: " + str(x - 10) + " y: " + str(y - 10))
            koda_om = row[koda_om_col - 1].value
            lokacija = row[lokacija_col - 1].value
            tip = row[tip_col - 1].value
            dimenzija = row[dimenzija_col - 1].value
            naslov = row[naslov_col - 1].value
            lastnik = row[lastnik_col - 1].value
            serijska = row[seriska_col - 1].value
            rf = row[rf_col - 1].value

            seznam_info.append(
                [
                    koda_om,
                    lastnik,
                    naslov,
                    tip,
                    dimenzija,
                    lokacija,
                    serijska,
                    rf_col,
                    x,
                    y,
                ]
            )

        # Print the list to verify the values
        seznam_infotim = [list(row) for row in seznam_info]
        print("Seznam:", seznam_infotim)
        print("Število vodomerov: " + str(len(seznam_infotim)))

        # Prenesi vodomere v ArcGis Pro
        stevec = 0

        fields = [
            "Koda_OM",
            "serijska_stevilka",
            "Lastnik",
            "Naslov",
            "Dimenzija",
            "GPS_Lat",
            "GPS_Lon",
            "Lokacija",
            "RF_Naslov",
            "Tip_oddajnika",
            "SHAPE@XY",
        ]
        edit = arcpy.da.Editor(env.workspace)
        edit.startEditing(False, True)
        edit.startOperation()

        sr = arcpy.Describe(vodomer_fc).spatialReference

        for row in seznam_infotim:
            stevec += 1
            print("Števec: " + str(stevec) + " " + row[1])
            try:
                with arcpy.da.InsertCursor(vodomer_fc, fields) as cursor:
                    new_row = (
                        row[0],  # om id
                        row[6],  # serijska
                        row[1],  # lastnik
                        row[2],  # naslov
                        row[4],  # dimenzija
                        row[9],  # lat
                        row[8],  # lon
                        row[5],  # lokacija
                        row[7],  # rf
                        row[3],  # tip
                        (row[9], row[8]),  # x, y
                    )
                    print("new row ", new_row)
                    cursor.insertRow(new_row)
            except Exception as e:
                # self.napaka(e)
                self.logc("Napaka pri dodajanju vodomerov " + str(e), err)
            finally:
                self.progress.emit(int(stevec / rows * 100))
                self.status.emit(f"{stevec} / {rows}")

        edit.stopOperation()
        edit.stopEditing(True)
        # Update status & progress bar
        # self.progress.emit(int((stevec + 1) / rows * 100))
        # self.status.emit(f"{stevec + 1} / {rows}")

        # workbook.save(filename)

        self.presledek()
        minutes, seconds = stop(start_time)

        self.logc(f"Končano v {minutes} min {seconds} sek.", blue)
        # msgbox = MsgBox("Končano")

    def log(self, in_str):
        self.logger.izpisi(in_str, black)

    def logc(self, in_str, barva):
        self.logger.izpisi(in_str, barva)

    def presledek(self):
        self.log("\n")
